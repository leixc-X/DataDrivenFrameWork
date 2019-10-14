# encoding=utf-8
import openpyxl
from openpyxl.styles import Border, Side, Font
import time
"""
本文件解析Excel文件的方法封装
"""


class ParseExcel(object):

    def __init__(self):
        self.workbook = None
        self.excelFile = None
        # 设置字体颜色
        self.font = Font(color=None)
        # 颜色对应的RGB值
        self.RGBDict = {'red': 'FFFF3030', 'green': 'FF008B00'}

    def loadWorkBook(self, excelPathAndName):
        # 将Excel文件加载到内存，并获取workbook对象
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile = excelPathAndName
        return self.workbook

    def getSheetByName(self, sheetName):
        # 根据sheet名获取sheet对象
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self, sheetIndex):
        # 根据sheet的索引号获取sheet对象
        try:
            sheetname = self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e
        sheet = self.workbook.get_sheet_by_name(sheetname)
        return sheet

    def getRowsNumber(self, sheet):
        # 获取sheet中有数据区域的结束行号
        return sheet.max_row

    def getColsNumber(self, sheet):
        # 获取sheet中有数据区域的结束列号
        return sheet.max_column

    def getStartRowNumber(self, sheet):
        # 获取sheet中有数据区域的开始行号
        return sheet.min_row

    def getStartColsNumber(self, sheet):
        # 获取sheet中有数据区域的开始列号
        return sheet.min_column

    def getRow(self, sheet, rowNo):
        # 获取sheet中某一行，返回的是这一行所有内容组成的tuple，
        # 从下标1开始，sheet.row[1]表示第一行
        try:
            return sheet.rows[rowNo - 1]
        except Exception as e:
            raise e

    def getColumn(self, sheet, colNo):
        # 获取sheet中某一列，返回的是这一列所有内容组成的tuple，
        # 从下标1开始，sheet.columns[1]表示第一列
        try:
            return sheet.columns[colNo - 1]
        except Exception as e:
            raise e

    def getCellOfValue(self, sheet, coordinate = None, rowNo = None, colsNo = None):
        # 根据单元所在位置索引获取单元格中的值，下标从1开始
        # sheet.cell(row =1, column = 1).value表示Excel中第一行第一列的值
        if coordinate != None:
            try:
                return sheet.cell(coordinate=coordinate).value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Corrdinates of cell!")

    def getCellOfObject(self, sheet, coordinate = None, rowNo = None, colsNo = None):
        # 获取某个单元格对象，可以根据单元格所在位置的数字索引，
        # 也可以直接根据Excel中单元格的编码以及坐标
        # 如 getCellObject(sheet, coordinate = 'AI') or
        # getCellObject(sheet, rowNow =1, colsNo =2)
        if coordinate != None:
            try:
                return sheet.cell(coordinate=coordinate)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Corrdinates of cell!")

    def writeCell(self, sheet, content, coordinate=None, rowNo=None, colsNo=None, style = None):
        # 根据单元格在Excel中的编码坐标或者数字索引坐标向单元格写入数据，
        # 下标从1开始，参数style表示字体颜色的名字，如red，green
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate = coordinate).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)

            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value = content
                if style:
                    sheet.cell(row = rowNo, column=colsNo).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Corrdinates of cell!")

    def writCellCurrentTime(self, sheet, coordinate=None, rowNo=None, colsNo=None):
        # 写入当前的时间，下标从1开始
        now = int(time.time()) # 显示为时间戳
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = currentTime
                self.workbook.save(self.excelFile)

            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                sheet.cell(row=rowNo, column=colsNo).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Corrdinates of cell!")

if __name__ == '__main__':
    # 测试代码
    pe = ParseExcel()
    # 测试所用的Excel文件“126邮箱联系人.xlsx”需要手动创建
    pe.loadWorkBook(r"C:\Users\Mr雷的电脑\Desktop\DataDrivenFrameWork\testData\126邮箱联系人.xlsx")
    print("通过名称获取sheet对象名字：", pe.getSheetByName(u"联系人").title)
    print(pe.getSheetByIndex(0).title)
    sheet = pe.getSheetByIndex(0)
    print(type(sheet))
    print(pe.getRowsNumber(sheet))
    print(pe.getColsNumber(sheet))
    rows = pe.getRow(sheet, 1)
    for i in list(rows):
        print(i.value)
    print(pe.getCellOfObject(sheet, rowNo=1, colsNo=1))
    # pe.writeCell(sheet, u"我爱我的祖国", rowNo=10, colsNo=10)
    pe.writCellCurrentTime(sheet, rowNo=10, colsNo=11)